from openpyxl import Workbook, load_workbook
import os

# Initialize services and staff files
def initialize_files():
    # Services file
    if not os.path.exists('services.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Services"
        ws.append(["Service Name", "Category", "Duration", "Price"])  # Columns
        wb.save('services.xlsx')

    # Staff file
    if not os.path.exists('staff.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff"
        ws.append(["Staff Name", "Role", "Email"])  # Columns
        wb.save('staff.xlsx')

    # Bookings file
    if not os.path.exists('bookings.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(["Booking ID", "Customer Name", "Service Name", "Date", "Time", "Staff Name", "Status"])  # Columns
        wb.save('bookings.xlsx')

# Fetch all staff
def get_staff():
    initialize_files()
    wb = load_workbook('staff.xlsx')
    ws = wb.active
    staff = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        staff.append(row)
    return staff

# Add a new staff member
def add_staff(name, role, email):
    initialize_files()
    wb = load_workbook('staff.xlsx')
    ws = wb.active
    ws.append([name, role, email])
    wb.save('staff.xlsx')

# Delete a staff member
def delete_staff(row_index):
    initialize_files()
    wb = load_workbook('staff.xlsx')
    ws = wb.active
    ws.delete_rows(row_index + 1)  # Excel index starts from 1
    wb.save('staff.xlsx')

# Fetch all bookings
def get_all_bookings():
    initialize_files()
    wb = load_workbook('bookings.xlsx')
    ws = wb.active
    bookings = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        bookings.append(row)
    return bookings

# Cancel a booking
def cancel_booking(row_index):
    initialize_files()
    wb = load_workbook('bookings.xlsx')
    ws = wb.active
    ws.cell(row_index + 1, 7, "Cancelled")  # Update status to "Cancelled"
    wb.save('bookings.xlsx')
