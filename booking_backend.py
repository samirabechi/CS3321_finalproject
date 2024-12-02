from openpyxl import load_workbook, Workbook
import os
import uuid

# Initialize bookings file
def initialize_bookings():
    if not os.path.exists('bookings.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(["Booking ID", "Customer Name", "Service Name", "Date", "Time", "Staff Name", "Status"])  # Columns
        wb.save('bookings.xlsx')

# Fetch available services
def get_services():
    return ["Lash Extensions", "Brow Tinting", "Lash Lift", "Facial Treatment"]  # Hardcoded for now

# Save a new booking
def save_booking(customer_name, service_name, date, time, staff_name=""):
    initialize_bookings()
    wb = load_workbook('bookings.xlsx')
    ws = wb["Bookings"]

    # Generate a unique booking ID
    booking_id = str(uuid.uuid4())[:8]

    # Save the booking to the Excel file
    ws.append([booking_id, customer_name, service_name, date, time, staff_name, "Upcoming"])
    wb.save('bookings.xlsx')
    return booking_id

# Fetch booking history for a customer
def get_booking_history(customer_name):
    initialize_bookings()
    wb = load_workbook('bookings.xlsx')
    ws = wb["Bookings"]

    history = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == customer_name:  # Match customer name
            history.append(row)
    return history

# Cancel a booking
def cancel_booking(booking_id):
    initialize_bookings()
    wb = load_workbook('bookings.xlsx')
    ws = wb["Bookings"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == booking_id:  # Match booking ID
            row[6].value = "Cancelled"  # Update status to "Cancelled"
            wb.save('bookings.xlsx')
            return True
    return False
def get_staff_names():
    # Return a hardcoded list of staff for now
    return ["Alice", "Bob", "Charlie", "Diana", "Eve"]

def get_services():
    # Return a hardcoded list of services with categories for now
    return {
        "Lash Lift": ["Lash Lift - 1h", "Lash Lift & Tint - 1h"],
        "Brow Services": ["Brow Tint - 20min"],
        "Face Waxing": ["Lip Wax - 5min", "Eyebrows Wax - 10min"],
        "Eyelash Extensions": ["Classic Lash Full Set - 2h", "Mega Volume Lash Full Set - 2h 15min"]
    }
