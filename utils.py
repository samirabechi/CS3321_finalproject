import os
from openpyxl import Workbook, load_workbook
import bcrypt

# Initialize the Excel file if it doesn't exist
def initialize_excel():
    if not os.path.exists('users.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Password", "Role"])  # Column headers
        wb.save('users.xlsx')

# Save new user data to Excel
def save_user(username, password, role):
    wb = load_workbook('users.xlsx')
    ws = wb["Users"]
    ws.append([username, password, role])
    wb.save('users.xlsx')

# Check user credentials during login
def check_credentials(username, password):
    wb = load_workbook('users.xlsx')
    ws = wb['Users']

    for row in ws.iter_rows(min_row=2, values_only=True):  # Start from the second row
        stored_username, stored_password, role = row
        if stored_username == username and bcrypt.checkpw(password.encode(), stored_password.encode()):
            return role  # Return the role if credentials match
    return None  # Return None if credentials are invalid
