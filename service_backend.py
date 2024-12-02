from openpyxl import Workbook, load_workbook
import os

# Initialize services file
def initialize_services():
    if not os.path.exists('services.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Services"
        ws.append(["Service Name", "Category", "Duration", "Price"])  # Columns
        wb.save('services.xlsx')

# Get all services
def get_services():
    initialize_services()
    wb = load_workbook('services.xlsx')
    ws = wb["Services"]

    services = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        services.append(row)
    return services

# Add a new service
def add_service(service_name, category, duration, price):
    initialize_services()
    wb = load_workbook('services.xlsx')
    ws = wb["Services"]
    ws.append([service_name, category, duration, price])
    wb.save('services.xlsx')

# Edit an existing service
def edit_service(row_index, service_name, category, duration, price):
    initialize_services()
    wb = load_workbook('services.xlsx')
    ws = wb["Services"]
    ws.delete_rows(row_index + 1)  # Delete old entry (Excel index starts from 1)
    ws.insert_rows(row_index + 1)
    ws.cell(row_index + 1, 1, service_name)
    ws.cell(row_index + 1, 2, category)
    ws.cell(row_index + 1, 3, duration)
    ws.cell(row_index + 1, 4, price)
    wb.save('services.xlsx')

# Delete a service
def delete_service(row_index):
    initialize_services()
    wb = load_workbook('services.xlsx')
    ws = wb["Services"]
    ws.delete_rows(row_index + 1)  # Excel index starts from 1
    wb.save('services.xlsx')
